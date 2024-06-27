import re
import openpyxl

def parse_header_file(file_path):
    with open(file_path, 'r') as file:
        content = file.read()

    # Regular expression to match function prototypes
    pattern = r'([\w\s]+)\s+(\w+)\s*\(([\w\s,]*)\)\s*;'
    prototypes = re.findall(pattern, content)

    return prototypes

def create_excel_sheet(prototypes):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write column headers
    worksheet['A1'] = 'ID'
    worksheet['B1'] = 'Function Name'
    worksheet['C1'] = 'Return Type'
    worksheet['D1'] = 'Parameters'

    # Write function prototypes to the Excel sheet
    row = 2
    for prototype in prototypes:
        return_type, function_name, parameters = prototype
        worksheet.cell(row=row, column=1, value=f'IDX{row-2}')
        worksheet.cell(row=row, column=2, value=function_name)
        worksheet.cell(row=row, column=3, value=return_type)
        worksheet.cell(row=row, column=4, value=parameters)
        row += 1

    workbook.save('function_prototypes.xlsx')

if __name__ == "__main__":
    header_file_path = '/home/saidmagdy/Desktop/EL_dip24_Tasks/python_tasks/extract_funPrototype_Cfile/C_Header_File.h'
    prototypes = parse_header_file(header_file_path)
    create_excel_sheet(prototypes)
    print("Function prototypes saved to 'function_prototypes.xlsx'")